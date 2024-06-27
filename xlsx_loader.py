from openpyxl import Workbook, load_workbook


def index_to_cell(index: int) -> str:
    chars: list[str] = [
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
    ]
    return chars[index % 8] + str(index // 8 + 1)


# workbook: Workbook = load_workbook(filename="template.xlsx")
# worksheet = workbook.active

# header, *rows = [row for row in worksheet.iter_rows(values_only=True)]

# data: list[list[str | None]] = [[j for j in i[1:]] for i in rows]
# data: list[str | None] = [item for sublist in data for item in sublist]

# pick_up_from: list[str] = [
#     index_to_cell(i) for i in range(len(data)) if data[i] is not None
# ]

# print(pick_up_from)


def update_template_file(
    pick_up_from: list[str], template_file_path: str, output_file_path: str
):
    with open(template_file_path, "r") as file:
        template_code = file.read()

    pick_up_from_assignment = f"self.pick_up_from: list[str] = {a}"
    updated_code = template_code.replace(
        "self.pick_up_from: list[str] = []", pick_up_from_assignment
    )

    with open(output_file_path, "w") as file:
        file.write(updated_code)


a = ["A1", "B1", "C1"]
template_file_path = "templates/template.py"
output_file_path = "updated_template.py"
update_template_file(a, template_file_path, output_file_path)
